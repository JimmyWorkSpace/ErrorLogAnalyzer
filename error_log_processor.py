import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook

def process_error_log(input_path: str, existing_excel_path: str, sheet_date: str):
    df = pd.read_excel(input_path)

    required_cols = {'ERROR_MESSAGE', 'START_TIME', 'END_TIME', 'ERROR_LEVEL', 'ERROR_CODE'}
    if not required_cols.issubset(df.columns):
        missing = required_cols - set(df.columns)
        raise ValueError(f"Missing columns in input file: {missing}")

    df['START_TIME'] = pd.to_datetime(df['START_TIME'], errors='coerce')
    df['END_TIME'] = pd.to_datetime(df['END_TIME'], errors='coerce')

    filtered_df = df[
        (df['ERROR_LEVEL'] == 'Alarm') &
        (df['ERROR_CODE'].astype(str).str.startswith('1')) &
        (df['ERROR_CODE'].astype(str) != '1001')
    ].copy()

    output_rows = []
    grouped = filtered_df.groupby('ERROR_MESSAGE')
    total_count = 0

    for message, group in grouped:
        group_sorted = group.sort_values(by='START_TIME')
        count = len(group_sorted)
        total_count += count

        output_rows.append({
            'ERROR_MESSAGE': message,
            'START_TIME': group_sorted['START_TIME'].min(),
            'END_TIME': group_sorted['END_TIME'].max(),
            'Count': count
        })

        for _, row in group_sorted.iterrows():
            output_rows.append({
                'ERROR_MESSAGE': '',
                'START_TIME': row['START_TIME'],
                'END_TIME': row['END_TIME'],
                'Count': ''
            })

    book = load_workbook(existing_excel_path)
    util_sheet_name = f"{sheet_date}_Utilization"

    if util_sheet_name not in book.sheetnames:
        raise ValueError(f"❌ Sheet '{util_sheet_name}' not found.")

    util_df = pd.read_excel(existing_excel_path, sheet_name=util_sheet_name)
    used_ohts = util_df.loc[0, 'Used OHTs']
    failure_rate = total_count / (used_ohts * 24)
    failure_rate_str = f"{failure_rate:.2%}"

    output_rows.append({
        'ERROR_MESSAGE': 'Failure Rate',
        'START_TIME': '',
        'END_TIME': '',
        'Count': failure_rate_str
    })

    result_df = pd.DataFrame(output_rows)

    error_sheet_name = f"{sheet_date}_errorStatistics"
    with pd.ExcelWriter(existing_excel_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        result_df.to_excel(writer, sheet_name=error_sheet_name, index=False)

    print(f"✅ Exported to: {existing_excel_path} (sheet: {error_sheet_name})")

    weekly_df = pd.read_excel(existing_excel_path, sheet_name="Weekly_Report")
    target_date = datetime.strptime(sheet_date, "%Y%m%d").date()

    updated = False
    for i, row in weekly_df.iterrows():
        try:
            row_date = pd.to_datetime(row['Date'], errors='coerce').date()
        except:
            continue
        if row_date == target_date:
            weekly_df.at[i, 'Failure Rate (%)'] = failure_rate
            updated = True
            print(f"📌 Inserted Failure Rate ({failure_rate_str}) into Weekly_Report at row {i+2}")
            break

    if not updated:
        print("⚠️ Could not find matching date row in Weekly_Report to insert Failure Rate")

    # Recompute "Weekly Avg" row
    df_data = weekly_df[weekly_df['Date'] != 'Weekly Avg'].copy()
    weekly_avg_row = {
        'Date': 'Weekly Avg',
        'Avg EXECUTE PERIOD': pd.to_numeric(df_data['Avg EXECUTE PERIOD'], errors='coerce').mean(),
        'OHT Utilization (%)': df_data['OHT Utilization (%)']
            .dropna().apply(lambda x: float(str(x).replace('%', ''))).mean(),
        'Failure Rate (%)': df_data['Failure Rate (%)']
            .dropna().apply(lambda x: float(str(x).replace('%', ''))).mean(),
        'Transfer Count': df_data['Transfer Count'].sum()
    }

    weekly_avg_row['OHT Utilization (%)'] = f"{weekly_avg_row['OHT Utilization (%)']:.2f}%"
    weekly_avg_row['Failure Rate (%)'] = f"{weekly_avg_row['Failure Rate (%)']:.5f}%"

    # Remove old avg row and append new one
    weekly_df = weekly_df[weekly_df['Date'] != 'Weekly Avg']
    weekly_df = pd.concat([weekly_df, pd.DataFrame([weekly_avg_row])], ignore_index=True)

    # Save back Weekly_Report
    with pd.ExcelWriter(existing_excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        weekly_df.to_excel(writer, sheet_name='Weekly_Report', index=False)

    # Move Weekly_Report to end
    wb = load_workbook(existing_excel_path)
    if 'Weekly_Report' in wb.sheetnames:
        ws = wb['Weekly_Report']
        wb._sheets.remove(ws)
        wb._sheets.append(ws)
        wb.save(existing_excel_path)
        print("📦 Weekly_Report updated and moved to the end.")
    wb.close()
